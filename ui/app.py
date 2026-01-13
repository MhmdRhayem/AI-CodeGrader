import gradio as gr
import requests
import json
import logging
from typing import Tuple, Optional, Dict, Any, List

import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

API_URL = "http://localhost:8000/api/v1/grade"
BATCH_API_URL = "http://localhost:8000/api/v1/grade_batch"
REQUEST_TIMEOUT = 120

# ---------------- Sample data ----------------
SAMPLE_PROBLEM = """Write a C++ function that finds and returns the maximum value in an integer array.

Function signature: int findMax(int arr[], int size);

Requirements:
- The function should find the largest element in the array
- Array size is guaranteed to be at least 1
- Handle both positive and negative numbers correctly"""

SAMPLE_REFERENCE = """int findMax(int arr[], int size) {
    int max = arr[0];
    for (int i = 1; i < size; i++) {
        if (arr[i] > max) {
            max = arr[i];
        }
    }
    return max;
}"""

SAMPLE_RUBRIC = """{
  "total_points": 10,
  "criteria": [
    {
      "name": "Correctness",
      "description": "Does the function correctly find and return the maximum value?",
      "max_points": 5.0,
      "evaluation_guidelines": "Test with positive numbers, negative numbers, and mixed values."
    },
    {
      "name": "Code Quality",
      "description": "Is the code clean, readable, and well-structured?",
      "max_points": 3.0,
      "evaluation_guidelines": "Check for clear variable names, consistent formatting, and readable logic."
    },
    {
      "name": "Efficiency",
      "description": "Is the solution optimal in terms of time and space complexity?",
      "max_points": 2.0,
      "evaluation_guidelines": "Time should be O(n), space should be O(1)."
    }
  ]
}"""

SAMPLE_STUDENT = """int findMax(int arr[], int size) {
    int max = arr[0];
    for (int i = 1; i < size; i++) {
        if (arr[i] > max) {
            max = arr[i];
        }
    }
    return max;
}"""


# ---------------- Helpers ----------------
def _parse_and_validate_rubric(rubric_json: str) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    if not rubric_json.strip():
        return None, "Rubric cannot be empty."
    try:
        rubric = json.loads(rubric_json)
    except json.JSONDecodeError as e:
        return None, f"Invalid rubric JSON: {str(e)}"

    if not isinstance(rubric, dict):
        return None, "Rubric must be a JSON object."
    if "criteria" not in rubric or "total_points" not in rubric:
        return None, "Rubric must include 'criteria' and 'total_points'."

    criteria = rubric.get("criteria", [])
    if not isinstance(criteria, list) or len(criteria) == 0:
        return None, "'criteria' must be a non-empty list."

    criteria_sum = sum(float(c.get("max_points", 0) or 0) for c in criteria)
    total_points = float(rubric.get("total_points", 0) or 0)

    if abs(criteria_sum - total_points) > 0.01:
        return None, (
            "Rubric mismatch:\n"
            f"- Sum of criteria max_points: {criteria_sum:.1f}\n"
            f"- total_points: {total_points:.1f}\n"
            "These must be equal."
        )

    return rubric, None


def _format_score(result: Dict[str, Any]) -> str:
    final_score = float(result.get("final_score", 0) or 0)
    total_points = float(result.get("total_points", 0) or 0)
    percentage = float(result.get("percentage", 0) or 0)

    if percentage >= 90:
        icon, message = "🌟", "Excellent work!"
    elif percentage >= 80:
        icon, message = "✨", "Great job!"
    elif percentage >= 70:
        icon, message = "👍", "Good work!"
    elif percentage >= 60:
        icon, message = "📚", "Needs improvement"
    else:
        icon, message = "⚠️", "Requires attention"

    return (
        f"## {icon} Final Score\n\n"
        f"**{final_score:.1f} / {total_points:.1f}**  \n"
        f"**{percentage:.1f}%**\n\n"
        f"{message}"
    )


def _breakdown_table(result: Dict[str, Any]) -> List[List[Any]]:
    breakdown = result.get("breakdown", []) or []
    rows: List[List[Any]] = []
    for c in breakdown:
        name = c.get("criterion_name", "Unknown")
        awarded = float(c.get("points_awarded", 0) or 0)
        max_pts = float(c.get("max_points", 0) or 0)
        feedback = (c.get("feedback", "") or "").strip()
        pct = (awarded / max_pts * 100.0) if max_pts > 0 else 0.0
        rows.append([name, round(awarded, 2), round(max_pts, 2), f"{pct:.0f}%", feedback])
    return rows


def _format_overall_feedback(result: Dict[str, Any]) -> str:
    overall = (result.get("overall_feedback", "") or "").strip()
    return overall if overall else "_No overall feedback provided._"


def _format_meta(result: Dict[str, Any]) -> str:
    strategy_names = {
        "cot": "Chain-of-Thought",
        "few_shot_cot": "Few-Shot CoT",
        "voting": "Voting Consensus",
        "evaluator_optimizer": "Evaluator-Optimizer",
    }
    s = strategy_names.get(result.get("grading_strategy", "unknown"), "Unknown")
    m = result.get("model_used", "N/A")
    return f"**Strategy:** {s}  \n**Model:** {m}"


def _read_uploaded_cpp(files) -> List[Dict[str, str]]:
    subs: List[Dict[str, str]] = []
    if not files:
        return subs

    for f in files:
        path = getattr(f, "name", None) or f  # supports multiple gradio versions
        filename = os.path.basename(str(path))
        with open(path, "r", encoding="utf-8", errors="ignore") as fp:
            code = fp.read()
        subs.append({"filename": filename, "student_code": code})
    return subs


def _make_batch_excel_report(data: Dict[str, Any], rubric_total: float) -> str:
    wb = Workbook()

    # ---------------- Sheet 1: Summary ----------------
    ws1 = wb.active
    ws1.title = "Summary"

    headers1 = ["Filename", "Final Score", "Total Points", "Percentage", "Status", "Error"]
    ws1.append(headers1)

    for col in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    results = data.get("results", []) or []
    for item in results:
        fname = item.get("filename", "unknown.cpp")
        if item.get("error"):
            ws1.append([fname, 0, rubric_total, 0, "FAILED", str(item.get("error"))])
        else:
            res = item.get("result", {}) or {}
            final_score = float(res.get("final_score", 0) or 0)
            total_points = float(res.get("total_points", rubric_total) or rubric_total)
            percentage = float(res.get("percentage", 0) or 0)
            ws1.append([fname, final_score, total_points, percentage, "OK", ""])

    # ---------------- Sheet 2: Criterion Breakdown ----------------
    ws2 = wb.create_sheet("Breakdown")
    headers2 = ["Filename", "Criterion", "Awarded", "Max", "Feedback"]
    ws2.append(headers2)

    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for item in results:
        fname = item.get("filename", "unknown.cpp")
        if item.get("error"):
            continue
        res = item.get("result", {}) or {}
        breakdown = res.get("breakdown", []) or []
        for c in breakdown:
            ws2.append([
                fname,
                c.get("criterion_name", ""),
                float(c.get("points_awarded", 0) or 0),
                float(c.get("max_points", 0) or 0),
                (c.get("feedback", "") or "").strip(),
            ])

    # ---------------- Sheet 3: Feedback + Reasoning ----------------
    ws3 = wb.create_sheet("Feedback")
    headers3 = ["Filename", "Overall Feedback", "Reasoning Trace", "Strategy", "Model"]
    ws3.append(headers3)

    for col in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    MAX_REASONING_CHARS = 12000
    for item in results:
        fname = item.get("filename", "unknown.cpp")
        if item.get("error"):
            ws3.append([fname, "", "", "", ""])
            continue

        res = item.get("result", {}) or {}
        overall = (res.get("overall_feedback", "") or "").strip()
        reasoning = (res.get("reasoning_trace", "") or "").strip()
        if len(reasoning) > MAX_REASONING_CHARS:
            reasoning = reasoning[:MAX_REASONING_CHARS] + "\n\n...[truncated]..."

        ws3.append([
            fname,
            overall,
            reasoning,
            res.get("grading_strategy", ""),
            res.get("model_used", ""),
        ])

    # ---------------- Autosize columns ----------------
    def autosize(ws, max_width=80):
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, max_width)

    autosize(ws1, max_width=60)
    autosize(ws2, max_width=70)
    autosize(ws3, max_width=80)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(tempfile.gettempdir(), f"batch_grading_results_{ts}.xlsx")
    wb.save(out_path)
    return out_path


def _processing_placeholder():
    md = "# ⏳ Grading…\n\n**Please wait while we evaluate the submission.**"
    breakdown_placeholder = [["⏳ Grading…", "", "", "", "Please wait while we evaluate the submission."]]
    return md, breakdown_placeholder, md, "", gr.update(value=None, visible=False)


def grade_submission(
    problem_desc: str,
    reference_sol: str,
    rubric_json: str,
    student_code: str,
    student_files,
    mode: str,
    strategy: str,
):
    # Validate text fields
    if not problem_desc.strip():
        return "❌ Problem description cannot be empty.", [], "", "", gr.update(value=None, visible=False)
    if not reference_sol.strip():
        return "❌ Reference solution cannot be empty.", [], "", "", gr.update(value=None, visible=False)

    rubric, rubric_err = _parse_and_validate_rubric(rubric_json)
    if rubric_err:
        return f"❌ {rubric_err}", [], "", "", gr.update(value=None, visible=False)

    # ---------------- Batch mode ----------------
    if mode == "batch":
        submissions = _read_uploaded_cpp(student_files)
        if not submissions:
            return "❌ Please upload at least one .cpp file for batch grading.", [], "", "", gr.update(value=None, visible=False)

        payload = {
            "problem_description": problem_desc,
            "reference_solution": reference_sol,
            "rubric": rubric,
            "grading_strategy": strategy,
            "submissions": submissions,
        }

        try:
            r = requests.post(BATCH_API_URL, json=payload, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            data = r.json()

            rows: List[List[Any]] = []
            total_points = float(rubric.get("total_points", 0) or 0)

            for item in data.get("results", []):
                fname = item.get("filename", "unknown.cpp")
                if item.get("error"):
                    rows.append([fname, 0, total_points, "0%", "❌ " + str(item.get("error"))])
                else:
                    res = item.get("result", {}) or {}
                    rows.append([
                        fname,
                        float(res.get("final_score", 0) or 0),
                        float(res.get("total_points", total_points) or total_points),
                        f"{float(res.get('percentage', 0) or 0):.0f}%",
                        "✅ graded",
                    ])

            summary_md = (
                "# ✅ Batch grading complete\n\n"
                f"**Files:** {data.get('count', len(rows))}  \n"
                f"**Succeeded:** {data.get('ok', 0)}  \n"
                f"**Failed:** {data.get('count', len(rows)) - data.get('ok', 0)}"
            )

            feedback_md = "Batch grading finished."
            reasoning_md = ""

            report_path = _make_batch_excel_report(data, total_points)
            report_update = gr.update(value=report_path, visible=True)  # clickable filename download

            return summary_md, rows, feedback_md, reasoning_md, report_update

        except requests.exceptions.ConnectionError:
            return (
                "❌ Connection Error\n\nCannot connect to the batch grading API.\n\n"
                "**Check:** backend is running and reachable at `http://localhost:8000`.",
                [],
                "—",
                "",
                gr.update(value=None, visible=False),
            )
        except requests.exceptions.Timeout:
            return (
                "❌ Request Timeout\n\nThe request exceeded the configured timeout.\n\n"
                "Try a faster strategy (Chain-of-Thought / Few-Shot CoT) or check backend load.",
                [],
                "—",
                "",
                gr.update(value=None, visible=False),
            )
        except requests.exceptions.HTTPError as e:
            try:
                detail = e.response.json().get("detail", e.response.text)
            except Exception:
                detail = e.response.text
            return f"❌ API Error\n\n{detail}", [], "—", "", gr.update(value=None, visible=False)
        except Exception as e:
            logger.error(f"Unexpected batch error: {str(e)}", exc_info=True)
            return f"❌ Unexpected error: {str(e)}", [], "—", "", gr.update(value=None, visible=False)

    # ---------------- Single mode ----------------
    if not student_code.strip():
        return "❌ Student code cannot be empty.", [], "", "", gr.update(value=None, visible=False)

    payload = {
        "problem_description": problem_desc,
        "reference_solution": reference_sol,
        "rubric": rubric,
        "student_code": student_code,
        "grading_strategy": strategy,
    }

    try:
        response = requests.post(API_URL, json=payload, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        result = response.json()

        score_md = _format_score(result)
        breakdown_rows = _breakdown_table(result)
        overall_md = _format_overall_feedback(result)

        reasoning = (result.get("reasoning_trace", "") or "").strip()
        meta_md = _format_meta(result)

        reasoning_md = ""
        if reasoning:
            reasoning_md = "## 🔍 Detailed Reasoning\n\n" + reasoning
        if meta_md:
            reasoning_md += ("\n\n---\n\n" + meta_md) if reasoning_md else meta_md

        return score_md, breakdown_rows, overall_md, reasoning_md, gr.update(value=None, visible=False)

    except requests.exceptions.ConnectionError:
        return (
            "❌ Connection Error\n\nCannot connect to the grading API.\n\n"
            "**Check:** backend is running and reachable at `http://localhost:8000`.",
            [],
            "—",
            "",
            gr.update(value=None, visible=False),
        )
    except requests.exceptions.Timeout:
        return (
            "❌ Request Timeout\n\nThe request exceeded the configured timeout.\n\n"
            "Try a faster strategy (Chain-of-Thought / Few-Shot CoT) or check backend load.",
            [],
            "—",
            "",
            gr.update(value=None, visible=False),
        )
    except requests.exceptions.HTTPError as e:
        try:
            detail = e.response.json().get("detail", e.response.text)
        except Exception:
            detail = e.response.text
        return f"❌ API Error\n\n{detail}", [], "—", "", gr.update(value=None, visible=False)
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}", exc_info=True)
        return f"❌ Unexpected error: {str(e)}", [], "—", "", gr.update(value=None, visible=False)


# ---------------- Theme (no CSS) ----------------
theme = (
    gr.themes.Soft(
        primary_hue="blue",
        secondary_hue="purple",
        neutral_hue="gray",
        font=[gr.themes.GoogleFont("Inter"), "sans-serif"],
    )
    .set(
        # Light
        body_background_fill="*neutral_50",
        block_background_fill="white",
        input_background_fill="white",
        block_border_color="*neutral_200",
        block_border_width="1px",
        block_radius="10px",
        shadow_drop="*shadow_md",
        # Dark
        body_background_fill_dark="*neutral_950",
        block_background_fill_dark="*neutral_900",
        input_background_fill_dark="*neutral_900",
        block_border_color_dark="*neutral_700",
        body_text_color_dark="*neutral_50",
    )
)

with gr.Blocks(title="C++ Grading Agent") as demo:
    gr.Markdown("# 🎓 C++ Grading Agent\nPaste inputs → choose strategy → grade.")

    with gr.Row():
        with gr.Column(scale=1):
            gr.Markdown("## 📝 Inputs")

            with gr.Tabs():
                with gr.Tab("Problem"):
                    problem_input = gr.TextArea(
                        label="Problem Description",
                        value=SAMPLE_PROBLEM,
                        lines=7,
                    )
                    reference_input = gr.Code(
                        label="Reference Solution (C++)",
                        language="cpp",
                        value=SAMPLE_REFERENCE,
                        lines=10,
                    )

                with gr.Tab("Rubric"):
                    rubric_input = gr.Code(
                        label="Grading Rubric (JSON)",
                        language="json",
                        value=SAMPLE_RUBRIC,
                        lines=14,
                    )
                    gr.Markdown("Tip: `total_points` must equal the sum of all `max_points`.")

                with gr.Tab("Student"):
                    mode_input = gr.Radio(
                        label="Submission Mode",
                        choices=[
                            ("Single (paste code)", "single"),
                            ("Batch (upload .cpp files)", "batch"),
                        ],
                        value="single",
                    )

                    student_input = gr.Code(
                        label="Student Submission (C++)",
                        language="cpp",
                        value=SAMPLE_STUDENT,
                        lines=12,
                        visible=True,
                    )

                    student_files = gr.Files(
                        label="Upload .cpp files (batch grading)",
                        file_types=[".cpp"],
                        file_count="multiple",
                        visible=False,
                    )

                    def _toggle_mode(mode):
                        return (
                            gr.update(visible=(mode == "single")),
                            gr.update(visible=(mode == "batch")),
                        )

                    mode_input.change(
                        fn=_toggle_mode,
                        inputs=[mode_input],
                        outputs=[student_input, student_files],
                        show_progress="minimal",
                    )

            gr.Markdown("## ⚙️ Controls")
            strategy_input = gr.Radio(
                label="Grading Strategy",
                choices=[
                    ("⚡ Chain-of-Thought", "cot"),
                    ("📚 Few-Shot CoT", "few_shot_cot"),
                    ("🗳️ Voting", "voting"),
                    ("🔄 Evaluator-Optimizer", "evaluator_optimizer"),
                ],
                value="cot",
            )

            grade_btn = gr.Button("🚀 Grade", variant="primary")

        with gr.Column(scale=1):
            gr.Markdown("## 📊 Results")

            with gr.Tabs():
                with gr.Tab("Summary"):
                    score_output = gr.Markdown(value="")

                with gr.Tab("Breakdown"):
                    breakdown_output = gr.Dataframe(
                        headers=["Criterion / File", "Awarded", "Max", "Percent", "Feedback / Status"],
                        datatype=["str", "number", "number", "str", "str"],
                        wrap=True,
                        interactive=False,
                        row_count=(0, "dynamic"),
                        column_count=(5, "fixed"),
                    )

                with gr.Tab("Feedback"):
                    overall_feedback_output = gr.Markdown(value="")
                    reasoning_output = gr.Markdown(value="")

            # OUTSIDE tabs: clickable filename download appears here only after batch success
            report_file = gr.File(
                label="📄 Batch report (.xlsx)",
                interactive=False,
                visible=False,
            )

    grade_btn.click(
        fn=_processing_placeholder,
        inputs=[],
        outputs=[score_output, breakdown_output, overall_feedback_output, reasoning_output, report_file],
        show_progress="minimal",
        queue=False,
    ).then(
        fn=grade_submission,
        inputs=[problem_input, reference_input, rubric_input, student_input, student_files, mode_input, strategy_input],
        outputs=[score_output, breakdown_output, overall_feedback_output, reasoning_output, report_file],
        show_progress="minimal",
    )

if __name__ == "__main__":
    logger.info("Starting Gradio UI on http://localhost:7860")
    demo.launch(
        server_name="127.0.0.1",
        server_port=7860,
        share=False,
        show_error=True,
        theme=theme,
    )
